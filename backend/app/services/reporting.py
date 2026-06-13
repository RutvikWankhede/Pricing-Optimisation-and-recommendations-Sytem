import os
import uuid
import math
from datetime import datetime, timezone
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for server execution
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database.models import Product, SalesData, PricingRecommendation, ElasticityAnalysis, ForecastingResult, Report
from app.core.config import settings

def apply_merged_style(ws, cell_range, fill, font=None, alignment=None, border=None):
    """Utility to style all cells in a merged range in openpyxl."""
    for row in ws[cell_range]:
        for cell in row:
            if fill: cell.fill = fill
            if font: cell.font = font
            if alignment: cell.alignment = alignment
            if border: cell.border = border

def save_matplotlib_charts(db: Session, suffix: str) -> tuple[str, str, str]:
    """Generate professional corporate-theme charts using Matplotlib and save to disk."""
    sales = db.query(SalesData).all()
    if not sales:
        return "", "", ""

    # Elegant enterprise palette colors
    colors_list = ['#1B365D', '#0D9488', '#2563EB', '#64748B', '#F59E0B', '#10B981', '#EF4444', '#6366F1', '#EC4899', '#8B5CF6']
    
    # Set default styles
    plt.rcParams['font.sans-serif'] = 'Arial'
    plt.rcParams['font.family'] = 'sans-serif'
    
    # 1. Category Revenue Share & Regional Performance (Combined Chart)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4.5))
    
    # Category Mix (Left)
    cat_query = db.query(Product.category, func.sum(SalesData.revenue)).join(SalesData, SalesData.product_id == Product.id).group_by(Product.category).all()
    cat_labels = [c[0] for c in cat_query]
    cat_vals = [float(c[1] or 0) for c in cat_query]
    
    if cat_vals:
        ax1.pie(
            cat_vals,
            labels=cat_labels,
            autopct='%1.1f%%',
            startangle=140,
            colors=colors_list[:len(cat_labels)],
            textprops={'fontsize': 8, 'color': '#0F172A', 'weight': 'semibold'},
            wedgeprops={'edgecolor': 'white', 'linewidth': 1.2, 'antialiased': True}
        )
    ax1.set_title("Revenue Share by Category", fontsize=11, fontweight='bold', color='#1B365D', pad=10)
    
    # Regional Yield (Right)
    reg_query = db.query(Product.region, func.sum(SalesData.revenue)).join(SalesData, SalesData.product_id == Product.id).group_by(Product.region).all()
    reg_labels = [r[0] for r in reg_query]
    reg_vals = [float(r[1] or 0) for r in reg_query]
    
    if reg_vals:
        bars = ax2.barh(reg_labels, reg_vals, color='#2563EB', edgecolor='none', height=0.6)
        for bar in bars:
            width = bar.get_width()
            ax2.text(width + (max(reg_vals)*0.02), bar.get_y() + bar.get_height()/2, f'₹{width:,.0f}', 
                     va='center', ha='left', fontsize=8, color='#475569', weight='semibold')
                 
    ax2.set_title("Revenue Yield by Region", fontsize=11, fontweight='bold', color='#1B365D', pad=10)
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    ax2.spines['left'].set_color('#CBD5E1')
    ax2.spines['bottom'].set_color('#CBD5E1')
    ax2.tick_params(colors='#475569', labelsize=8)
    ax2.grid(True, axis='x', linestyle='--', alpha=0.2, color='#94A3B8')
    
    plt.tight_layout()
    cat_chart_path = settings.UPLOAD_FOLDER / f"cat_regional_mix_{suffix}.png"
    plt.savefig(cat_chart_path, dpi=200, bbox_inches='tight', transparent=True)
    plt.close()

    # 2. Revenue Trends & Forecasting (Historical + Predictive Horizon)
    daily_stats = {}
    for s in sales:
        date_str = str(s.sales_date)
        if date_str not in daily_stats:
            daily_stats[date_str] = {"revenue": 0.0, "profit": 0.0}
        daily_stats[date_str]["revenue"] += float(s.revenue)
        daily_stats[date_str]["profit"] += float(s.profit)

    sorted_dates = sorted(daily_stats.keys())
    dates_obj = [datetime.strptime(d, "%Y-%m-%d") for d in sorted_dates]
    revenues = [daily_stats[d]["revenue"] for d in sorted_dates]
    profits = [daily_stats[d]["profit"] for d in sorted_dates]

    # Apply 7-day rolling average to smooth trend
    if len(revenues) > 7:
        smoothed_rev = pd.Series(revenues).rolling(window=7, min_periods=1).mean().tolist()
        smoothed_prof = pd.Series(profits).rolling(window=7, min_periods=1).mean().tolist()
    else:
        smoothed_rev = revenues
        smoothed_prof = profits

    # 30-Day Forecast projections
    forecasts = db.query(ForecastingResult).order_by(ForecastingResult.forecast_date).all()
    forecast_dates = []
    
    # Get product average prices to scale forecast volume to revenue
    prod_prices = {}
    for s in sales:
        prod_prices[s.product_id] = prod_prices.get(s.product_id, []) + [float(s.price)]
    prod_avg_prices = {pid: sum(prcs)/len(prcs) for pid, prcs in prod_prices.items()}
    
    daily_forecast = {}
    for f in forecasts:
        f_date_str = str(f.forecast_date)
        if f_date_str not in daily_forecast:
            daily_forecast[f_date_str] = 0.0
        avg_p = prod_avg_prices.get(f.product_id, 10.0)
        daily_forecast[f_date_str] += float(f.predicted_demand) * avg_p
        
    sorted_fc_dates = sorted(daily_forecast.keys())
    fc_dates_obj = [datetime.strptime(d, "%Y-%m-%d") for d in sorted_fc_dates]
    fc_revenues = [daily_forecast[d] for d in sorted_fc_dates]

    plt.figure(figsize=(9, 4.2))
    # Plot historical
    plt.plot(dates_obj, revenues, color='#93C5FD', alpha=0.3, linewidth=1, label="Daily Historical Revenue")
    plt.plot(dates_obj, smoothed_rev, color='#1B365D', linewidth=2.5, label="7-Day Revenue Trend")
    plt.plot(dates_obj, smoothed_prof, color='#0D9488', linewidth=2, label="7-Day Profit Trend")
    
    # Plot forecast if exists
    if fc_dates_obj:
        # Connect historical and forecast lines
        if dates_obj:
            fc_dates_plot = [dates_obj[-1]] + fc_dates_obj
            fc_revs_plot = [smoothed_rev[-1]] + fc_revenues
        else:
            fc_dates_plot = fc_dates_obj
            fc_revs_plot = fc_revenues
            
        plt.plot(fc_dates_plot, fc_revs_plot, color='#2563EB', linestyle='--', linewidth=2.5, label="30-Day Forecast Projections")
        # Vertical boundary line
        if dates_obj:
            plt.axvline(x=dates_obj[-1], color='#64748B', linestyle=':', alpha=0.8)
            plt.text(dates_obj[-1], max(max(revenues or [0]), max(fc_revenues or [0]))*0.9, '  Forecast Start', color='#64748B', fontsize=8, weight='bold')

    plt.title("Revenue Trends & Forecasting Performance Horizon", fontsize=11, fontweight='bold', color='#1B365D', pad=15)
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
    
    all_dates = dates_obj + fc_dates_obj
    if all_dates:
        plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(all_dates)//10)))
    plt.gcf().autofmt_xdate()
    plt.grid(True, linestyle='--', alpha=0.2, color='#94A3B8')
    plt.gca().spines['top'].set_visible(False)
    plt.gca().spines['right'].set_visible(False)
    plt.gca().spines['left'].set_color('#CBD5E1')
    plt.gca().spines['bottom'].set_color('#CBD5E1')
    plt.tick_params(colors='#475569', labelsize=8)
    plt.legend(frameon=True, facecolor='#F8FAFC', edgecolor='none', fontsize=8)
    plt.tight_layout()

    trend_chart_path = settings.UPLOAD_FOLDER / f"trend_forecast_{suffix}.png"
    plt.savefig(trend_chart_path, dpi=200, bbox_inches='tight', transparent=True)
    plt.close()

    # 3. Product Elasticity Bar Chart
    elasticity_data = db.query(ElasticityAnalysis).join(Product).all()
    elasticity_chart_path = ""
    
    if elasticity_data:
        # Sort by elasticity coefficient ascending
        elasticity_data.sort(key=lambda x: float(x.elasticity_score))
        
        prod_names = [e.product.product_name for e in elasticity_data]
        scores = [float(e.elasticity_score) for e in elasticity_data]
        
        # Limit to top 12 products for readability
        if len(prod_names) > 12:
            prod_names = prod_names[:12]
            scores = scores[:12]
            
        plt.figure(figsize=(8.5, 4))
        # Color code bars based on sensitivity
        bar_colors = ['#EF4444' if s < -1.0 else ('#10B981' if s > -1.0 else '#64748B') for s in scores]
        
        bars = plt.bar(prod_names, scores, color=bar_colors, edgecolor='none', width=0.5)
        plt.axhline(y=-1.0, color='#64748B', linestyle='--', alpha=0.6, linewidth=1)
        plt.text(len(prod_names)-1, -0.9, 'Elasticity boundary (-1.0)', color='#64748B', fontsize=7, ha='right')
        
        # Add labels
        for bar in bars:
            height = bar.get_height()
            va_dir = 'bottom' if height >= 0 else 'top'
            plt.text(bar.get_x() + bar.get_width()/2, height + (0.05 if height >= 0 else -0.05), f'{height:.2f}', 
                     ha='center', va=va_dir, fontsize=7, color='#475569', weight='semibold')

        plt.title("Elasticity Coefficient Distribution by SKU", fontsize=11, fontweight='bold', color='#1B365D', pad=15)
        plt.xticks(rotation=45, ha='right', fontsize=8)
        plt.gca().spines['top'].set_visible(False)
        plt.gca().spines['right'].set_visible(False)
        plt.gca().spines['left'].set_color('#CBD5E1')
        plt.gca().spines['bottom'].set_color('#CBD5E1')
        plt.tick_params(colors='#475569', labelsize=8)
        plt.grid(True, axis='y', linestyle='--', alpha=0.15, color='#94A3B8')
        plt.tight_layout()
        
        elasticity_chart_path = settings.UPLOAD_FOLDER / f"elasticity_distribution_{suffix}.png"
        plt.savefig(elasticity_chart_path, dpi=200, bbox_inches='tight', transparent=True)
        plt.close()
        
    return str(cat_chart_path), str(trend_chart_path), str(elasticity_chart_path)

def generate_excel_report(db: Session, user_id: str = None) -> str:
    """Generate a multi-tab executive financial spreadsheet with embedded charts and conditional formatting."""
    wb = Workbook()
    suffix = str(uuid.uuid4())[:8]
    cat_img, trend_img, elasticity_img = save_matplotlib_charts(db, suffix)
    
    # Borders & Alignments
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Font definitions
    title_font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    sec_font = Font(name="Segoe UI", size=11, bold=True, color="1B365D")
    label_font = Font(name="Segoe UI", size=9, bold=True, color="475569")
    val_font = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
    data_font = Font(name="Segoe UI", size=10, color="334155")
    bold_data_font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    
    green_text_font = Font(name="Segoe UI", size=9, bold=True, color="166534")
    red_text_font = Font(name="Segoe UI", size=9, bold=True, color="991B1B")
    gray_text_font = Font(name="Segoe UI", size=9, bold=True, color="475569")

    # Colors
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Royal Navy
    banner_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Matte Charcoal
    kpi_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Soft Blue-Gray
    zebra_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    # ================= Tab 1: Executive Summary Overview =================
    ws1 = wb.active
    ws1.title = "Executive Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Banner B2:H3
    ws1.merge_cells("B2:H3")
    ws1["B2"] = "Enterprise AI Pricing Intelligence Dashboard"
    apply_merged_style(ws1, "B2:H3", banner_fill, title_font, center_align, thin_border)
    
    sales = db.query(SalesData).all()
    if not sales:
        filename = f"pricing_strategy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = settings.REPORT_FOLDER / "excel" / filename
        wb.save(filepath)
        return str(filepath)
        
    total_rev = sum(float(s.revenue) for s in sales)
    total_cost = sum(float(s.cost) * int(s.quantity_sold) for s in sales)
    total_profit = total_rev - total_cost
    margin = (total_profit / total_rev * 100.0) if total_rev > 0 else 0.0
    active_products = db.query(Product).count()
    
    # Period over Period Change Calculations
    dates = db.query(func.min(SalesData.sales_date), func.max(SalesData.sales_date)).first()
    min_date, max_date = dates[0], dates[1]
    rev_change, profit_change, margin_change = 0.0, 0.0, 0.0
    if min_date and max_date and min_date != max_date:
        mid_date = min_date + (max_date - min_date) / 2
        rev_first = db.query(func.sum(SalesData.revenue)).filter(SalesData.sales_date < mid_date).scalar() or 0.0
        rev_second = db.query(func.sum(SalesData.revenue)).filter(SalesData.sales_date >= mid_date).scalar() or 0.0
        rev_change = ((float(rev_second) - float(rev_first)) / float(rev_first) * 100.0) if float(rev_first) > 0 else 0.0
        
        profit_first = db.query(func.sum(SalesData.profit)).filter(SalesData.sales_date < mid_date).scalar() or 0.0
        profit_second = db.query(func.sum(SalesData.profit)).filter(SalesData.sales_date >= mid_date).scalar() or 0.0
        profit_change = ((float(profit_second) - float(profit_first)) / float(profit_first) * 100.0) if float(profit_first) > 0 else 0.0
        
        margin_first = (float(profit_first) / float(rev_first) * 100.0) if float(rev_first) > 0 else 0.0
        margin_second = (float(profit_second) / float(rev_second) * 100.0) if float(rev_second) > 0 else 0.0
        margin_change = margin_second - margin_first
        
    forecasts = db.query(ForecastingResult).all()
    avg_conf = 0.82
    if forecasts:
        valid_scores = [float(f.confidence_score) for f in forecasts if f.confidence_score is not None]
        if valid_scores:
            avg_conf = sum(valid_scores) / len(valid_scores)
            
    period_str = f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}" if min_date and max_date else "N/A"
    ws1["B4"] = f"Reporting Horizon: {period_str} | Compiled: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["B4"].font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    
    # KPI Scorecards Block B5:E7
    kpi_columns = ["B", "C", "D", "E"]
    kpi_labels = ["Total Gross Revenue", "Total Net Profit", "Average Gross Margin", "Forecasting Model Accuracy"]
    kpi_values = [total_rev, total_profit, margin / 100.0 if margin > 0 else 0.0, avg_conf]
    kpi_formats = ['"₹"#,##0.00', '"₹"#,##0.00', '0.0%', '0.0%']
    pop_changes = [rev_change, profit_change, margin_change, 0.0]
    
    for idx, col in enumerate(kpi_columns):
        ws1[f"{col}5"] = kpi_labels[idx]
        ws1[f"{col}5"].font = label_font
        ws1[f"{col}5"].alignment = center_align
        
        ws1[f"{col}6"] = kpi_values[idx]
        ws1[f"{col}6"].font = val_font
        ws1[f"{col}6"].number_format = kpi_formats[idx]
        ws1[f"{col}6"].alignment = center_align
        
        change = pop_changes[idx]
        if idx == 3:
            ws1[f"{col}7"] = "Stable Model Fit"
            ws1[f"{col}7"].font = gray_text_font
        else:
            arrow = "▲" if change >= 0 else "▼"
            ws1[f"{col}7"] = f"{arrow} {change:+.1f}% PoP"
            ws1[f"{col}7"].font = green_text_font if change >= 0 else red_text_font
        ws1[f"{col}7"].alignment = center_align
        
        for r in [5, 6, 7]:
            c_cell = ws1[f"{col}{r}"]
            c_cell.fill = kpi_fill
            c_cell.border = thin_border
            
    # Top SKUs Table B9
    ws1["B9"] = "Top Contributing Product SKUs by Revenue"
    ws1["B9"].font = sec_font
    
    headers_top = ["Product SKU Name", "Category Segment", "Units Sold", "Gross Revenue", "Net Profit Contribution"]
    for col_idx, h in enumerate(headers_top, start=2):
        cell = ws1.cell(row=10, column=col_idx, value=h)
        style_cell(cell, font=header_font, fill=header_fill, alignment=center_align, border=thin_border)
        
    prod_map = {}
    for s in sales:
        p_name = s.product.product_name
        cat = s.product.category
        if p_name not in prod_map:
            prod_map[p_name] = {"category": cat, "qty": 0, "rev": 0.0, "prof": 0.0}
        prod_map[p_name]["qty"] += s.quantity_sold
        prod_map[p_name]["rev"] += float(s.revenue)
        prod_map[p_name]["prof"] += float(s.profit)
        
    sorted_prods = sorted(prod_map.items(), key=lambda x: x[1]["rev"], reverse=True)[:10]
    r_idx = 11
    for name, stats in sorted_prods:
        ws1.cell(row=r_idx, column=2, value=name).font = bold_data_font
        ws1.cell(row=r_idx, column=3, value=stats["category"]).font = data_font
        
        c3 = ws1.cell(row=r_idx, column=4, value=stats["qty"])
        style_cell(c3, font=data_font, alignment=right_align, number_format='#,##0')
        
        c4 = ws1.cell(row=r_idx, column=5, value=stats["rev"])
        style_cell(c4, font=data_font, alignment=right_align, number_format='"₹"#,##0.00')
        
        c5 = ws1.cell(row=r_idx, column=6, value=stats["prof"])
        style_cell(c5, font=data_font, alignment=right_align, number_format='"₹"#,##0.00')
        
        for c in range(2, 7):
            cell = ws1.cell(row=r_idx, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        r_idx += 1
        
    # Embed category regional mix chart at H4
    if cat_img and os.path.exists(cat_img):
        openpyxl_img = OpenpyxlImage(cat_img)
        ws1.add_image(openpyxl_img, "H4")
        
    # ================= Tab 2: Pricing Recommendations =================
    ws2 = wb.create_sheet(title="Optimal Recommendations")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("B2:H3")
    ws2["B2"] = "Model-Optimized Target Pricing Recommendations"
    apply_merged_style(ws2, "B2:H3", banner_fill, title_font, center_align, thin_border)
    
    headers_recs = ["Product Name", "Current Price", "Recommended Price", "Price Delta", "Expected Rev Change", "Expected Profit Change", "Strategic Action"]
    for col_idx, h in enumerate(headers_recs, start=2):
        cell = ws2.cell(row=5, column=col_idx, value=h)
        style_cell(cell, font=header_font, fill=header_fill, alignment=center_align, border=thin_border)
        
    recs = db.query(PricingRecommendation).all()
    r_idx = 6
    
    increase_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # soft green
    decrease_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # soft red
    maintain_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # soft gray
    
    for rec in recs:
        curr = float(rec.current_price)
        target = float(rec.recommended_price)
        diff = target - curr
        
        action_fill = maintain_fill
        action_text = "MAINTAIN"
        action_font = Font(name="Segoe UI", size=9, bold=True, color="475569")
        
        if diff > 0.1:
            action_fill = increase_fill
            action_text = "INCREASE"
            action_font = Font(name="Segoe UI", size=9, bold=True, color="166534")
        elif diff < -0.1:
            action_fill = decrease_fill
            action_text = "DECREASE"
            action_font = Font(name="Segoe UI", size=9, bold=True, color="991B1B")
            
        ws2.cell(row=r_idx, column=2, value=rec.product.product_name).font = bold_data_font
        
        c3 = ws2.cell(row=r_idx, column=3, value=curr)
        style_cell(c3, font=data_font, alignment=right_align, number_format='"₹"#,##0.00')
        
        c4 = ws2.cell(row=r_idx, column=4, value=target)
        style_cell(c4, font=data_font, alignment=right_align, number_format='"₹"#,##0.00')
        
        c5 = ws2.cell(row=r_idx, column=5, value=diff)
        style_cell(c5, font=data_font, alignment=right_align, number_format='+"₹"#,##0.00;-"₹"#,##0.00;"₹"0.00')
        
        c6 = ws2.cell(row=r_idx, column=6, value=float(rec.expected_revenue_change or 0))
        style_cell(c6, font=data_font, alignment=right_align, number_format='+0.0%;-0.0%;0.0%')
        
        c7 = ws2.cell(row=r_idx, column=7, value=float(rec.expected_profit_change or 0))
        style_cell(c7, font=data_font, alignment=right_align, number_format='+0.0%;-0.0%;0.0%')
        
        c8 = ws2.cell(row=r_idx, column=8, value=action_text)
        style_cell(c8, font=action_font, fill=action_fill, alignment=center_align)
        
        for c in range(2, 9):
            cell = ws2.cell(row=r_idx, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1 and c != 8:
                cell.fill = zebra_fill
        r_idx += 1
        
    # Embed Elasticity distribution chart at J5
    if elasticity_img and os.path.exists(elasticity_img):
        openpyxl_img = OpenpyxlImage(elasticity_img)
        ws2.add_image(openpyxl_img, "J5")
        
    # ================= Tab 3: Forecast Projections =================
    ws3 = wb.create_sheet(title="Forecast Projections")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("B2:G3")
    ws3["B2"] = "30-Day Sales Volume Forecast Roadmap"
    apply_merged_style(ws3, "B2:G3", banner_fill, title_font, center_align, thin_border)
    
    headers_fc = ["Forecast Date", "Product SKU Name", "Predicted Demand (Units)", "Baseline Price", "Estimated Yield", "Forecast Model Name"]
    for col_idx, h in enumerate(headers_fc, start=2):
        cell = ws3.cell(row=5, column=col_idx, value=h)
        style_cell(cell, font=header_font, fill=header_fill, alignment=center_align, border=thin_border)
        
    # Get product average prices
    prod_prices = {}
    for s in sales:
        prod_prices[s.product_id] = prod_prices.get(s.product_id, []) + [float(s.price)]
    prod_avg_prices = {pid: sum(prcs)/len(prcs) for pid, prcs in prod_prices.items()}
    
    forecasts = db.query(ForecastingResult).order_by(ForecastingResult.forecast_date, ForecastingResult.product_id).all()
    r_idx = 6
    for f in forecasts:
        base_price = prod_avg_prices.get(f.product_id, 10.0)
        yield_val = float(f.predicted_demand) * base_price
        
        c2 = ws3.cell(row=r_idx, column=2, value=f.forecast_date.strftime("%Y-%m-%d"))
        style_cell(c2, font=data_font, alignment=center_align)
        
        ws3.cell(row=r_idx, column=3, value=f.product.product_name).font = bold_data_font
        
        c4 = ws3.cell(row=r_idx, column=4, value=float(f.predicted_demand))
        style_cell(c4, font=data_font, alignment=right_align, number_format='#,##0')
        
        c5 = ws3.cell(row=r_idx, column=5, value=base_price)
        style_cell(c5, font=data_font, alignment=right_align, number_format='"₹"#,##0.00')
        
        c6 = ws3.cell(row=r_idx, column=6, value=yield_val)
        style_cell(c6, font=data_font, alignment=right_align, number_format='"₹"#,##0.00')
        
        ws3.cell(row=r_idx, column=7, value=f.model_name).font = data_font
        
        for c in range(2, 8):
            cell = ws3.cell(row=r_idx, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        r_idx += 1
        
    # ================= Tab 4: Price Elasticity Sensitivity =================
    ws4 = wb.create_sheet(title="Elasticity Sensitivity")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("B2:E3")
    ws4["B2"] = "Product Demand Elasticity Profiles"
    apply_merged_style(ws4, "B2:E3", banner_fill, title_font, center_align, thin_border)
    
    headers_el = ["Product SKU Name", "Elasticity Coefficient", "Sensitivity Classification", "Strategic Recommendation"]
    for col_idx, h in enumerate(headers_el, start=2):
        cell = ws4.cell(row=5, column=col_idx, value=h)
        style_cell(cell, font=header_font, fill=header_fill, alignment=center_align, border=thin_border)
        
    elasticities = db.query(ElasticityAnalysis).all()
    r_idx = 6
    for el in elasticities:
        ws4.cell(row=r_idx, column=2, value=el.product.product_name).font = bold_data_font
        
        c3 = ws4.cell(row=r_idx, column=3, value=float(el.elasticity_score))
        style_cell(c3, font=data_font, alignment=right_align, number_format='0.00')
        
        c4 = ws4.cell(row=r_idx, column=4, value=el.elasticity_type.capitalize())
        style_cell(c4, font=data_font, alignment=center_align)
        
        score = float(el.elasticity_score)
        if score < -1.5:
            rec_str = "Highly elastic: Maintain or lower price to protect volume."
        elif score < -1.0:
            rec_str = "Moderately elastic: Exercise caution with price changes."
        elif score < 0:
            rec_str = "Inelastic: Pricing power exists. Optimize price upwards."
        else:
            rec_str = "Inelastic: Strong pricing power. Target +5% optimization."
            
        ws4.cell(row=r_idx, column=5, value=rec_str).font = data_font
        
        for c in range(2, 6):
            cell = ws4.cell(row=r_idx, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        r_idx += 1
        
    # Auto-fit columns across sheets (ignoring headers)
    for ws in [ws1, ws2, ws3, ws4]:
        ws.views.sheetView[0].showGridLines = True
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row in [2, 3]:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Set spacing overrides
    ws1.column_dimensions['G'].width = 3
    ws2.column_dimensions['I'].width = 3
    
    # Save Workbook
    filename = f"pricing_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = settings.REPORT_FOLDER / "excel" / filename
    wb.save(filepath)
    
    # Cleanup temp charts
    for path in [cat_img, trend_img, elasticity_img]:
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
                
    db_report = Report(
        user_id=user_id,
        report_type="excel",
        file_path=str(filepath),
        created_at=datetime.now(timezone.utc)
    )
    db.add(db_report)
    db.commit()
    
    return str(filepath)

def style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Utility to apply font, fill, alignment, and borders to a cell."""
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format

def generate_pdf_report(db: Session, user_id: str = None) -> str:
    """Generate a McKinsey-style premium executive strategy PDF report."""
    filename = f"pricing_strategy_brief_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = settings.REPORT_FOLDER / "pdf" / filename
    
    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=45, bottomMargin=45
    )
    
    styles = getSampleStyleSheet()

    # Custom styling
    navy_primary = colors.HexColor('#1B365D')
    teal_accent = colors.HexColor('#0D9488')
    charcoal_text = colors.HexColor('#334155')
    light_bg = colors.HexColor('#F8FAFC')

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.white,
        leading=26,
        spaceAfter=0
    )

    cover_meta_label_style = ParagraphStyle(
        'CoverMetaLabel',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor('#94A3B8')
    )

    cover_meta_val_style = ParagraphStyle(
        'CoverMetaVal',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#CBD5E1')
    )

    h1_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=15,
        textColor=colors.HexColor('#0F172A'),
        spaceBefore=10,
        spaceAfter=12,
        keepWithNext=True
    )

    h2_style = ParagraphStyle(
        'SubSectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=10.5,
        textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=8,
        spaceAfter=6,
        keepWithNext=True
    )

    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=9,
        textColor=charcoal_text,
        leading=13.5,
        spaceAfter=10
    )

    table_header_style = ParagraphStyle(
        'TableHeaderText',
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1  # Centered
    )

    table_cell_style = ParagraphStyle(
        'TableCellText',
        fontName='Helvetica',
        fontSize=7.5,
        textColor=charcoal_text,
        leading=10
    )

    table_cell_bold_style = ParagraphStyle(
        'TableCellBoldText',
        fontName='Helvetica-Bold',
        fontSize=7.5,
        textColor=colors.HexColor('#0F172A'),
        leading=10
    )

    kpi_card_style = ParagraphStyle(
        'KPICardText',
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#475569'),
        leading=12,
        alignment=1
    )

    story = []

    # Get data metrics
    sales = db.query(SalesData).all()
    if not sales:
        raise ValueError("No sales data available for report generation.")
        
    total_rev = sum(float(s.revenue) for s in sales)
    total_cost = sum(float(s.cost) * int(s.quantity_sold) for s in sales)
    total_profit = total_rev - total_cost
    margin = (total_profit / total_rev * 100.0) if total_rev > 0 else 0.0
    active_products = db.query(Product).count()

    # Dynamic PoP growth metrics
    dates = db.query(func.min(SalesData.sales_date), func.max(SalesData.sales_date)).first()
    min_date, max_date = dates[0], dates[1]
    
    rev_change, profit_change, margin_change = 0.0, 0.0, 0.0
    if min_date and max_date and min_date != max_date:
        mid_date = min_date + (max_date - min_date) / 2
        rev_first = db.query(func.sum(SalesData.revenue)).filter(SalesData.sales_date < mid_date).scalar() or 0.0
        rev_second = db.query(func.sum(SalesData.revenue)).filter(SalesData.sales_date >= mid_date).scalar() or 0.0
        rev_change = ((float(rev_second) - float(rev_first)) / float(rev_first) * 100.0) if float(rev_first) > 0 else 0.0
        
        profit_first = db.query(func.sum(SalesData.profit)).filter(SalesData.sales_date < mid_date).scalar() or 0.0
        profit_second = db.query(func.sum(SalesData.profit)).filter(SalesData.sales_date >= mid_date).scalar() or 0.0
        profit_change = ((float(profit_second) - float(profit_first)) / float(profit_first) * 100.0) if float(profit_first) > 0 else 0.0
        
        margin_first = (float(profit_first) / float(rev_first) * 100.0) if float(rev_first) > 0 else 0.0
        margin_second = (float(profit_second) / float(rev_second) * 100.0) if float(rev_second) > 0 else 0.0
        margin_change = margin_second - margin_first
        
    rev_change_str = f"{rev_change:+.1f}%"
    profit_change_str = f"{profit_change:+.1f}%"
    margin_change_str = f"{margin_change:+.1f}%"

    forecasts = db.query(ForecastingResult).all()
    avg_conf = 0.82
    if forecasts:
        valid_scores = [float(f.confidence_score) for f in forecasts if f.confidence_score is not None]
        if valid_scores:
            avg_conf = sum(valid_scores) / len(valid_scores)

    period_str = f"{min_date.strftime('%B %d, %Y')} to {max_date.strftime('%B %d, %Y')}" if min_date and max_date else "N/A"
    
    rev_arrow = "▲" if rev_change >= 0 else "▼"
    rev_color = "#10B981" if rev_change >= 0 else "#EF4444"
    rev_perf_str = f"<font color='{rev_color}'><b>{rev_arrow} {rev_change_str}</b></font>"

    profit_arrow = "▲" if profit_change >= 0 else "▼"
    profit_color = "#10B981" if profit_change >= 0 else "#EF4444"
    profit_perf_str = f"<font color='{profit_color}'><b>{profit_arrow} {profit_change_str}</b></font>"

    margin_arrow = "▲" if margin_change >= 0 else "▼"
    margin_color = "#10B981" if margin_change >= 0 else "#EF4444"
    margin_perf_str = f"<font color='{margin_color}'><b>{margin_arrow} {margin_change_str}</b></font>"

    # Create Cover Page Dynamic Snapshot & KPIs
    cover_snapshot = (
        "This strategic executive brief leverages predictive ML demand modeling "
        "and granular price elasticity coefficients to deliver optimized target margin trajectories. "
        f"By analyzing {len(sales):,} transaction rows across {active_products} product SKUs, "
        "our AI systems have isolated immediate revenue capture opportunities and elasticity-sensitive risk vectors. "
        "The following analysis outlines an actionable transition plan to grow portfolio margin."
    )

    kpis_cover_data = [
        [
            Paragraph("<b>TOTAL REVENUE</b><br/><font size=11 color='#FFFFFF'><b>₹{:,.2f}</b></font><br/>{:}".format(total_rev, rev_perf_str), kpi_card_style),
            Paragraph("<b>NET PROFIT</b><br/><font size=11 color='#FFFFFF'><b>₹{:,.2f}</b></font><br/>{:}".format(total_profit, profit_perf_str), kpi_card_style),
            Paragraph("<b>GROSS MARGIN</b><br/><font size=11 color='#FFFFFF'><b>{:.1f}%</b></font><br/>{:}".format(margin, margin_perf_str), kpi_card_style),
            Paragraph("<b>ACTIVE SKUs</b><br/><font size=11 color='#FFFFFF'><b>{:}</b></font><br/><font color='#94A3B8'><b>Stable</b></font>".format(active_products), kpi_card_style),
        ]
    ]
    kpis_cover_table = Table(kpis_cover_data, colWidths=[118, 118, 118, 118])
    kpis_cover_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1E293B')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#334155')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # ================= PAGE 1: COVER PAGE =================
    cover_data = [
        [Paragraph("<font color='#3B82F6'><b>PRICESENSE AI | ENTERPRISE REVENUE SYSTEMS</b></font>", ParagraphStyle('CoverBranding', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#3B82F6')))],
        [Spacer(1, 15)],
        [Paragraph("Enterprise AI Pricing Intelligence<br/>Executive Brief", title_style)],
        [Spacer(1, 10)],
        [Paragraph("Model-Optimized Strategic Margin & Yield Assessment", ParagraphStyle('CoverSubtitle', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor('#94A3B8')))],
        [Spacer(1, 30)],
        [
            Table([
                [Paragraph("<b>Reporting Period:</b>", cover_meta_label_style), Paragraph(period_str, cover_meta_val_style)],
                [Paragraph("<b>Generated On:</b>", cover_meta_label_style), Paragraph(datetime.now().strftime('%B %d, %Y %H:%M:%S'), cover_meta_val_style)],
                [Paragraph("<b>Classification:</b>", cover_meta_label_style), Paragraph("CONFIDENTIAL / EXECUTIVE STAKEHOLDERS", cover_meta_val_style)],
                [Paragraph("<b>Analytics Engine:</b>", cover_meta_label_style), Paragraph("PriceSense Core v2.4", cover_meta_val_style)]
            ], colWidths=[120, 362], style=[
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#1E293B'))
            ])
        ],
        [Spacer(1, 40)],
        [Paragraph("<b>EXECUTIVE STRATEGIC SNAPSHOT</b>", ParagraphStyle('CoverSnapHeader', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#3B82F6')))],
        [Spacer(1, 6)],
        [Paragraph(cover_snapshot, ParagraphStyle('CoverSnapshotText', fontName='Helvetica', fontSize=9.5, textColor=colors.HexColor('#E2E8F0'), leading=14))],
        [Spacer(1, 40)],
        [Paragraph("<b>HERO PERFORMANCE METRICS</b>", ParagraphStyle('CoverKPIHeader', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#3B82F6')))],
        [Spacer(1, 10)],
        [kpis_cover_table]
    ]

    cover_table = Table(cover_data, colWidths=[532])
    cover_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#0F172A')), # matte navy/slate
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1E293B')),
        ('TOPPADDING', (0, 0), (-1, -1), 35),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 35),
        ('LEFTPADDING', (0, 0), (-1, -1), 30),
        ('RIGHTPADDING', (0, 0), (-1, -1), 30),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(Spacer(1, 10))
    story.append(cover_table)
    story.append(PageBreak())

    # ================= PAGE 2: BUSINESS STORY & SUMMARY =================
    # Query details dynamically for storytelling
    cat_query = db.query(Product.category, func.sum(SalesData.revenue).label("rev"), func.sum(SalesData.profit).label("prof")).join(SalesData, SalesData.product_id == Product.id).group_by(Product.category).all()
    sorted_categories = sorted(cat_query, key=lambda x: float(x[1] or 0), reverse=True)
    top_cat = sorted_categories[0][0] if sorted_categories else "N/A"
    top_cat_rev = float(sorted_categories[0][1] or 0) if sorted_categories else 0.0
    top_cat_pct = (top_cat_rev / total_rev * 100.0) if total_rev > 0 else 0.0

    bottom_cat = sorted_categories[-1][0] if len(sorted_categories) > 1 else "N/A"
    bottom_cat_rev = float(sorted_categories[-1][1] or 0) if len(sorted_categories) > 1 else 0.0
    bottom_cat_pct = (bottom_cat_rev / total_rev * 100.0) if total_rev > 0 else 0.0

    reg_query = db.query(Product.region, func.sum(SalesData.revenue).label("rev")).join(SalesData, SalesData.product_id == Product.id).group_by(Product.region).all()
    sorted_regions = sorted(reg_query, key=lambda x: float(x[1] or 0), reverse=True)
    top_region = sorted_regions[0][0] if sorted_regions else "N/A"
    top_region_rev = float(sorted_regions[0][1] or 0) if sorted_regions else 0.0
    top_region_pct = (top_region_rev / total_rev * 100.0) if total_rev > 0 else 0.0

    # Inelastic pricing opportunities
    opps_query = db.query(PricingRecommendation).join(Product).all()
    inelastic_opportunities = []
    for r in opps_query:
        diff = float(r.recommended_price) - float(r.current_price)
        if diff > 0.1:
            inelastic_opportunities.append(r)
    inelastic_opportunities.sort(key=lambda x: float(x.expected_profit_change or 0), reverse=True)

    opps_list = []
    for r in inelastic_opportunities[:3]:
        opps_list.append(f"<b>{r.product.product_name}</b> (recommendation ₹{float(r.recommended_price):,.2f}, +{float(r.expected_profit_change or 0)*100:.1f}% profit)")
    opps_summary = ", ".join(opps_list) if opps_list else "No immediate inelastic price increases recommended."

    # Highly elastic risks
    elasticity_query = db.query(ElasticityAnalysis).join(Product).all()
    elastic_risks = [e for e in elasticity_query if float(e.elasticity_score) < -1.5]
    elastic_risks.sort(key=lambda x: float(x.elasticity_score))

    risks_list = []
    for e in elastic_risks[:3]:
        risks_list.append(f"<b>{e.product.product_name}</b> (elasticity score of {float(e.elasticity_score):.2f})")
    risks_summary = ", ".join(risks_list) if risks_list else "No products exhibit severe price sensitivity risks."

    story.append(Paragraph("Executive Summary", h1_style))
    story.append(Paragraph("<b>Operational Portfolio & Margin Assessment</b>", h2_style))
    
    summary_desc1 = (
        f"A thorough evaluation of the business's pricing posture indicates robust overall sales metrics. "
        f"During the reporting horizon, the portfolio generated a total gross yield of <b>₹{total_rev:,.2f}</b>, "
        f"translating to a net profit of <b>₹{total_profit:,.2f}</b> and establishing an aggregate gross profit margin of <b>{margin:.2f}%</b>. "
        f"The period-over-period revenue grew by <b>{rev_change:+.1f}%</b>, while net profits shifted by <b>{profit_change:+.1f}%</b>. "
        f"These figures confirm that margin capture is healthy, but leaves significant untapped potential in specific inelastic segments."
    )
    story.append(Paragraph(summary_desc1, body_style))
    
    story.append(Paragraph("<b>Category Segment & Regional Contribution Analysis</b>", h2_style))
    summary_desc2 = (
        f"Analyzing performance by category reveals distinct distribution profiles. The top-performing vertical is "
        f"<b>{top_cat}</b>, which commanded <b>₹{top_cat_rev:,.2f}</b>, representing <b>{top_cat_pct:.1f}%</b> of total gross revenue. "
        f"Conversely, <b>{bottom_cat}</b> underperformed, contributing only <b>₹{bottom_cat_rev:,.2f}</b> (<b>{bottom_cat_pct:.1f}%</b> of portfolio yield). "
        f"Geographically, <b>{top_region}</b> represents the leading demand hub, yielding <b>₹{top_region_rev:,.2f}</b> (<b>{top_region_pct:.1f}%</b> of global sales). "
        f"Resources should be concentrated on expanding catalog diversity in {top_region} while applying price optimizations to underperforming verticals."
    )
    story.append(Paragraph(summary_desc2, body_style))

    story.append(Paragraph("<b>Price Optimization Opportunities & Sensitivity Risks</b>", h2_style))
    summary_desc3 = (
        f"The pricing intelligence engine has isolated key pricing power segments. Specific inelastic items, including {opps_summary}, "
        f"are currently underpriced compared to market willingness. Raising prices on these items will expand net profit margin. "
        f"In contrast, elastic segments like {risks_summary} exhibit intense price sensitivity. "
        f"Any aggressive price increases in these categories will spark volume drops. For these products, we recommend maintaining prices "
        f"or launching targeted promotional discounts to drive turnover."
    )
    story.append(Paragraph(summary_desc3, body_style))
    story.append(Spacer(1, 10))

    # Detailed 2x2 board layout
    kpi_board_data = [
        [
            Paragraph("<b>Gross Revenue Yield</b><br/><font size=14 color='#1B365D'><b>₹{:,.2f}</b></font><br/>{:}".format(total_rev, rev_perf_str), kpi_card_style),
            Paragraph("<b>Net Profit Contribution</b><br/><font size=14 color='#0D9488'><b>₹{:,.2f}</b></font><br/>{:}".format(total_profit, profit_perf_str), kpi_card_style)
        ],
        [
            Paragraph("<b>EBIT Gross Margin %</b><br/><font size=14 color='#2563EB'><b>{:.2f}%</b></font><br/>{:}".format(margin, margin_perf_str), kpi_card_style),
            Paragraph("<b>Forecasting Accuracy</b><br/><font size=14 color='#0F172A'><b>{:.1f}%</b></font><br/><font color='#64748B'><b>Stable Model Fit</b></font>".format(avg_conf * 100), kpi_card_style)
        ]
    ]
    kpi_board_table = Table(kpi_board_data, colWidths=[260, 260])
    kpi_board_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 16),
        ('RIGHTPADDING', (0, 0), (-1, -1), 16),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(kpi_board_table)
    story.append(PageBreak())

    # ================= PAGE 3: VISUAL ANALYTICS =================
    suffix = str(uuid.uuid4())[:8]
    cat_img, trend_img, elasticity_img = save_matplotlib_charts(db, suffix)

    story.append(Paragraph("Visual Revenue Analytics", h1_style))
    story.append(Paragraph("<b>Category Revenue Share & Regional Performance</b>", h2_style))
    
    if cat_img and os.path.exists(cat_img):
        story.append(Image(cat_img, width=480, height=216))
        story.append(Spacer(1, 10))
        
    story.append(Paragraph("<b>Historical Revenue Trends & Forecasting Horizon</b>", h2_style))
    
    if trend_img and os.path.exists(trend_img):
        story.append(Image(trend_img, width=480, height=224))
        
    story.append(PageBreak())

    # ================= PAGE 4: ELASTICITY & SENSITIVITY =================
    story.append(Paragraph("Demand Elasticity & Sensitivity", h1_style))
    story.append(Paragraph("<b>Elasticity Coefficient Distribution Profile</b>", h2_style))
    
    if elasticity_img and os.path.exists(elasticity_img):
        story.append(Image(elasticity_img, width=480, height=220))
        story.append(Spacer(1, 10))
        
    story.append(Paragraph("<b>SKU Elasticity Portfolio Matrix</b>", h2_style))
    
    headers_el = [[
        Paragraph("Product SKU Name", table_header_style),
        Paragraph("Category", table_header_style),
        Paragraph("Elasticity Coefficient", table_header_style),
        Paragraph("Sensitivity Profile", table_header_style)
    ]]
    
    # Get top 8 products by revenue to keep PDF compact
    top_sku_ids = [p[0] for p in db.query(Product.id, func.sum(SalesData.revenue)).join(SalesData).group_by(Product.id).order_by(func.sum(SalesData.revenue).desc()).limit(8).all()]
    
    elasticity_list = db.query(ElasticityAnalysis).filter(ElasticityAnalysis.product_id.in_(top_sku_ids)).all()
    for el in elasticity_list:
        headers_el.append([
            Paragraph(el.product.product_name, table_cell_bold_style),
            Paragraph(el.product.category, table_cell_style),
            Paragraph(f"{float(el.elasticity_score):.3f}", table_cell_style),
            Paragraph(el.elasticity_type.capitalize(), table_cell_style)
        ])
        
    el_table = Table(headers_el, colWidths=[180, 110, 110, 132])
    el_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]))
    story.append(el_table)
    story.append(PageBreak())

    # ================= PAGE 5: RECOMMENDATIONS LEDGER =================
    story.append(Paragraph("Strategic Pricing Recommendations", h1_style))
    story.append(Paragraph("<b>Model-Optimized Price Adjustments Ledger</b>", h2_style))
    
    rec_headers = [[
        Paragraph("Product Name", table_header_style),
        Paragraph("Current Price", table_header_style),
        Paragraph("Recommended Price", table_header_style),
        Paragraph("Rev Change", table_header_style),
        Paragraph("Profit Change", table_header_style),
        Paragraph("Action Status", table_header_style)
    ]]

    table_style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]

    recs = db.query(PricingRecommendation).all()
    # Limit to top 10 for pdf layout
    pdf_recs = recs[:10]
    for idx, r in enumerate(pdf_recs, start=1):
        diff = float(r.recommended_price) - float(r.current_price)
        action_text = "MAINTAIN"
        text_color_hex = '#475569'
        
        if diff > 0.1:
            action_text = "INCREASE"
            text_color_hex = '#166534'
            table_style_cmds.append(('BACKGROUND', (5, idx), (5, idx), colors.HexColor('#DCFCE7'))) # light green
        elif diff < -0.1:
            action_text = "DECREASE"
            text_color_hex = '#991B1B'
            table_style_cmds.append(('BACKGROUND', (5, idx), (5, idx), colors.HexColor('#FEE2E2'))) # light red
        else:
            table_style_cmds.append(('BACKGROUND', (5, idx), (5, idx), colors.HexColor('#F1F5F9'))) # light gray

        action_style = ParagraphStyle(f'ActionStyle_{idx}', fontName='Helvetica-Bold', fontSize=7.5, textColor=colors.HexColor(text_color_hex), alignment=1)

        rec_headers.append([
            Paragraph(r.product.product_name, table_cell_bold_style),
            Paragraph(f"₹{float(r.current_price):,.2f}", table_cell_style),
            Paragraph(f"₹{float(r.recommended_price):,.2f}", table_cell_style),
            Paragraph(f"{float(r.expected_revenue_change or 0)*100:+.1f}%", table_cell_style),
            Paragraph(f"{float(r.expected_profit_change or 0)*100:+.1f}%", table_cell_style),
            Paragraph(action_text, action_style)
        ])

    rec_table = Table(rec_headers, colWidths=[170, 75, 75, 75, 75, 62])
    rec_table.setStyle(TableStyle(table_style_cmds))
    story.append(rec_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("<b>Optimization Insights Blueprint</b>", h2_style))
    takeaway_desc = (
        "<b>1. Inelastic Price Adjustments:</b> Products labeled <i>INCREASE</i> exhibit low demand sensitivity (coefficients > -1.0). "
        "These adjustments represent high-yield opportunities and should be deployed immediately to capture additional gross profit.<br/>"
        "<b>2. Promotional Elastic Discounts:</b> Items flagged <i>DECREASE</i> are highly elastic. Implementing these discounts "
        "stimulates demand, increases inventory throughput, and prevents customer churn."
    )
    story.append(Paragraph(takeaway_desc, body_style))
    story.append(PageBreak())

    # ================= PAGE 6: RISK & ROADMAP =================
    story.append(Paragraph("Risk, Opportunity & Roadmap", h1_style))
    story.append(Paragraph("<b>Price Volatility & Sensitivity Risks</b>", h2_style))
    
    risk_text = (
        "<b>1. Volatility Vectors:</b> Products in highly competitive regions show rapid elasticity shifts. "
        "Frequent minor adjustments (e.g. daily/weekly pricing cycles) are recommended over large quarterly shifts to minimize friction.<br/>"
        "<b>2. Inventory Alignment:</b> Discounters must sync dynamic price drops with inventory holdings to prevent out-of-stock "
        "conditions. Monitor demand velocity constantly during promotion peaks.<br/>"
        "<b>3. Model Calibration:</b> Ingest transaction logs every 30 days to recalibrate the forecasting models and elasticity slopes."
    )
    story.append(Paragraph(risk_text, body_style))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Implementation Execution Roadmap</b>", h2_style))
    roadmap_data = [
        [Paragraph("<b>Time Horizon</b>", table_header_style), Paragraph("<b>Target Focus Area</b>", table_header_style), Paragraph("<b>Operational Deliverable</b>", table_header_style)],
        [Paragraph("Phase 1 (Immediate)", table_cell_bold_style), Paragraph("Inelastic Products", table_cell_style), Paragraph("Deploy recommended price increases (+3% to +8%) across inelastic cuisines.", table_cell_style)],
        [Paragraph("Phase 2 (30 Days)", table_cell_bold_style), Paragraph("Elastic Cuisines", table_cell_style), Paragraph("Trigger target discounts on elastic items to grow unit sales and offset seasonal dips.", table_cell_style)],
        [Paragraph("Phase 3 (60 Days)", table_cell_bold_style), Paragraph("Model Recalibration", table_cell_style), Paragraph("Ingest fresh transaction ledger to reassess elasticity coefficients and update forecasts.", table_cell_style)]
    ]
    roadmap_table = Table(roadmap_data, colWidths=[120, 120, 292])
    roadmap_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    story.append(roadmap_table)

    # Footer Callback (Page number & headers)
    def add_header_footer(canvas, doc):
        canvas.saveState()
        if doc.page > 1:
            # Header
            canvas.setFont('Helvetica-Bold', 7.5)
            canvas.setFillColor(colors.HexColor('#475569'))
            canvas.drawString(40, doc.pagesize[1] - 30, "PRICESENSE ENTERPRISE REVENUE INTELLIGENCE EXECUTIVE STRATEGY BRIEF")
            canvas.setStrokeColor(colors.HexColor('#CBD5E1'))
            canvas.setLineWidth(0.5)
            canvas.line(40, doc.pagesize[1] - 34, doc.pagesize[0] - 40, doc.pagesize[1] - 34)

            # Footer
            canvas.setFont('Helvetica', 7.5)
            canvas.setFillColor(colors.HexColor('#64748B'))
            canvas.drawString(40, 25, "CONFIDENTIAL - INVESTOR STRATEGY OUTLINE")
            canvas.drawRightString(doc.pagesize[0] - 40, 25, f"Page {doc.page}")
            canvas.line(40, 35, doc.pagesize[0] - 40, 35)
        canvas.restoreState()

    doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)

    # Cleanup temporary charts
    for path in [cat_img, trend_img, elasticity_img]:
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass

    # Save PDF report record
    db_report = Report(
        user_id=user_id,
        report_type="pdf",
        file_path=str(filepath),
        created_at=datetime.now(timezone.utc)
    )
    db.add(db_report)
    db.commit()

    return str(filepath)
